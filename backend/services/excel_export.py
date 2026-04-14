from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io


def generate_payroll_excel(payroll_data: list[dict], month: str) -> bytes:
    wb = Workbook()

    drivers = [p for p in payroll_data if p["person_type"] == "driver"]
    employees = [p for p in payroll_data if p["person_type"] == "employee"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = '#,##0.00'

    def _write_salary_sheet(ws, data: list[dict], title: str):
        ws.title = f"{title} - Salário"

        # Title row
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Folha de Salário - {title} - {month}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Nome", "CPF", "Salário Bruto (R$)", "INSS (R$)",
            "Adiantamento (R$)", "Salário Líquido (R$)", "Chave PIX",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_idx, p in enumerate(data, 4):
            breakdown = p.get("breakdown") or {}
            inss = float(p.get("inss") or 0)
            adv_totals = breakdown.get("advance_totals", {})
            # Only salary + product advances count for the salary sheet
            salary_adv = float(adv_totals.get("salario", 0)) + float(adv_totals.get("produtos", 0))
            gross = float(p["gross_pay"])
            net = round(gross - inss - salary_adv, 2)
            pix = breakdown.get("pix_key") or p.get("pix_key") or ""

            ws.cell(row=row_idx, column=1, value=p.get("person_name", "")).border = border
            ws.cell(row=row_idx, column=2, value=p.get("person_cpf", "")).border = border

            for col, val in [(3, gross), (4, inss), (5, salary_adv), (6, net)]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = money_fmt
                cell.border = border

            ws.cell(row=row_idx, column=7, value=pix).border = border

        # Column widths
        widths = {"A": 30, "B": 18, "C": 18, "D": 15, "E": 18, "F": 18, "G": 25}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        # Totals
        total_row = len(data) + 4
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        ws.cell(row=total_row, column=2).border = border

        for col, key_fn in [
            (3, lambda p: float(p["gross_pay"])),
            (4, lambda p: float(p.get("inss") or 0)),
            (5, lambda p: float((p.get("breakdown") or {}).get("advance_totals", {}).get("salario", 0)) +
                          float((p.get("breakdown") or {}).get("advance_totals", {}).get("produtos", 0))),
            (6, lambda p: float(p["gross_pay"]) - float(p.get("inss") or 0) -
                          float((p.get("breakdown") or {}).get("advance_totals", {}).get("salario", 0)) -
                          float((p.get("breakdown") or {}).get("advance_totals", {}).get("produtos", 0))),
        ]:
            cell = ws.cell(row=total_row, column=col, value=sum(key_fn(p) for p in data))
            cell.font = Font(bold=True)
            cell.number_format = money_fmt
            cell.border = border

        ws.cell(row=total_row, column=7).border = border

    def _write_benefit_sheet(ws, data: list[dict], title: str):
        ws.title = f"{title} - Benefício"

        ws.merge_cells("A1:I1")
        ws["A1"] = f"Folha de Benefícios - {title} - {month}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Nome", "Benefício Bruto (R$)",
            "Alimentação (R$)", "Dedução Alimentação (R$)",
            "Transporte (R$)", "Dedução Transporte (R$)",
            "Refeição (R$)", "Dedução Refeição (R$)",
            "Benefício Líquido (R$)",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        for row_idx, p in enumerate(data, 4):
            breakdown = p.get("breakdown") or {}
            benefit = breakdown.get("benefit", {})

            alim = float(p.get("beneficio_alimentacao") or benefit.get("alimentacao_valor", 0))
            trans = float(p.get("beneficio_transporte") or benefit.get("transporte_valor", 0))
            ref = float(p.get("beneficio_refeicao") or benefit.get("refeicao_valor", 0))
            gross_ben = alim + trans + ref

            ded_alim = float(benefit.get("alimentacao_deducao", 0))
            ded_trans = float(benefit.get("transporte_deducao", 0))
            ded_ref = float(benefit.get("refeicao_deducao", 0))
            net_ben = round(gross_ben - ded_alim - ded_trans - ded_ref, 2)

            ws.cell(row=row_idx, column=1, value=p.get("person_name", "")).border = border

            for col, val in [
                (2, gross_ben), (3, alim), (4, ded_alim),
                (5, trans), (6, ded_trans),
                (7, ref), (8, ded_ref), (9, net_ben),
            ]:
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.number_format = money_fmt
                cell.border = border

        # Column widths
        ws.column_dimensions["A"].width = 30
        for letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[letter].width = 20

        # Totals
        total_row = len(data) + 4
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border

        for col_idx in range(2, 10):
            vals = []
            for p in data:
                breakdown = p.get("breakdown") or {}
                benefit = breakdown.get("benefit", {})
                alim = float(p.get("beneficio_alimentacao") or benefit.get("alimentacao_valor", 0))
                trans = float(p.get("beneficio_transporte") or benefit.get("transporte_valor", 0))
                ref = float(p.get("beneficio_refeicao") or benefit.get("refeicao_valor", 0))
                ded_alim = float(benefit.get("alimentacao_deducao", 0))
                ded_trans = float(benefit.get("transporte_deducao", 0))
                ded_ref = float(benefit.get("refeicao_deducao", 0))
                row_vals = [
                    alim + trans + ref, alim, ded_alim,
                    trans, ded_trans, ref, ded_ref,
                    (alim + trans + ref) - (ded_alim + ded_trans + ded_ref),
                ]
                vals.append(row_vals[col_idx - 2])

            cell = ws.cell(row=total_row, column=col_idx, value=sum(vals))
            cell.font = Font(bold=True)
            cell.number_format = money_fmt
            cell.border = border

    # --- Build sheets ---

    # Sheet 1: Motoristas - Salário
    ws1 = wb.active
    if drivers:
        _write_salary_sheet(ws1, drivers, "Motoristas")
    else:
        ws1.title = "Motoristas - Salário"

    # Sheet 2: Motoristas - Benefício
    ws2 = wb.create_sheet()
    if drivers:
        _write_benefit_sheet(ws2, drivers, "Motoristas")
    else:
        ws2.title = "Motoristas - Benefício"

    # Sheet 3: Funcionários - Salário
    ws3 = wb.create_sheet()
    if employees:
        _write_salary_sheet(ws3, employees, "Funcionários")
    else:
        ws3.title = "Funcionários - Salário"

    # Sheet 4: Funcionários - Benefício
    ws4 = wb.create_sheet()
    if employees:
        _write_benefit_sheet(ws4, employees, "Funcionários")
    else:
        ws4.title = "Funcionários - Benefício"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
